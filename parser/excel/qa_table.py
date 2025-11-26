import logging
import random
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from excel_parser import RAGFlowExcelParser as Excel


def is_english(texts):
    """简单的英文检测函数"""
    if not texts:
        return False
    english_chars = sum(1 for text in texts for char in text if char.isalpha() and ord(char) < 128)
    total_chars = sum(1 for text in texts for char in text if char.isalpha())
    return english_chars / max(total_chars, 1) > 0.7


def rmPrefix(text):
    """移除常见前缀"""
    text = text.strip()
    prefixes = ['Q:', '问：', '问题：', 'A:', '答：', '答案：']
    for prefix in prefixes:
        if text.startswith(prefix):
            return text[len(prefix):].strip()
    return text


def random_choices(items, k=1):
    """随机选择k个元素"""
    return random.sample(items, min(k, len(items)))


class ExcelParser:
    """基础Excel解析器类"""
    def __init__(self):
        self.is_english = False


class QAExcel(ExcelParser):
    def __call__(self, fnm, binary=None, callback=None):
        if not binary:
            wb = load_workbook(fnm)
        else:
            wb = load_workbook(BytesIO(binary))
        total = 0
        for sheetname in wb.sheetnames:
            total += len(list(wb[sheetname].rows))

        res, fails = [], []
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            rows = list(ws.rows)
            for i, r in enumerate(rows):
                q, a = "", ""
                for cell in r:
                    if not cell.value:
                        continue
                    if not q:
                        q = str(cell.value)
                    elif not a:
                        a = str(cell.value)
                    else:
                        break
                if q and a:
                    res.append((q, a))
                else:
                    fails.append(str(i + 1))
                if len(res) % 999 == 0:
                    callback(len(res) *
                             0.6 /
                             total, ("Extract pairs: {}".format(len(res)) +
                                     (f"{len(fails)} failure, line: %s..." %
                                      (",".join(fails[:3])) if fails else "")))

        callback(0.6, ("Extract pairs: {}. ".format(len(res)) + (
            f"{len(fails)} failure, line: %s..." % (",".join(fails[:3])) if fails else "")))
        self.is_english = is_english(
            [rmPrefix(q) for q, _ in random_choices(res, k=30) if len(q) > 1])
        return res


class TableExcel(ExcelParser):
    def __call__(self, fnm, binary=None, from_page=0, to_page=10000000000, callback=None):
        if not binary:
            wb = Excel._load_excel_to_workbook(fnm)
        else:
            wb = Excel._load_excel_to_workbook(BytesIO(binary))
        total = 0
        for sheetname in wb.sheetnames:
            total += len(list(wb[sheetname].rows))
        res, fails, done = [], [], 0
        rn = 0
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            try:
                rows = list(ws.rows)
            except Exception as e:
                logging.warning(f"Skip sheet '{sheetname}' due to rows access error: {e}")
                continue
            if not rows:
                continue
            headers, header_rows = self._parse_headers(ws, rows)
            if not headers:
                continue
            data = []
            for i, r in enumerate(rows[header_rows:]):
                rn += 1
                if rn - 1 < from_page:
                    continue
                if rn - 1 >= to_page:
                    break
                row_data = self._extract_row_data(ws, r, header_rows + i, len(headers))
                if row_data is None:
                    fails.append(str(i))
                    continue
                if self._is_empty_row(row_data):
                    continue
                data.append(row_data)
                done += 1
            if len(data) == 0:
                continue
            df = pd.DataFrame(data, columns=headers)
            res.append(df)
        callback(0.3, ("Extract records: {}~{}".format(from_page + 1, min(to_page, from_page + rn)) + (
            f"{len(fails)} failure, line: %s..." % (",".join(fails[:3])) if fails else "")))
        return res

    def _parse_headers(self, ws, rows):
        if len(rows) == 0:
            return [], 0
        has_complex_structure = self._has_complex_header_structure(ws, rows)
        if has_complex_structure:
            return self._parse_multi_level_headers(ws, rows)
        else:
            return self._parse_simple_headers(rows)

    def _has_complex_header_structure(self, ws, rows):
        if len(rows) < 1:
            return False
        merged_ranges = list(ws.merged_cells.ranges)
        # 检查前两行是否涉及合并单元格
        for rng in merged_ranges:
            if rng.min_row <= 2:  # 只要合并区域涉及第1或第2行
                return True
        return False

    def _row_looks_like_header(self, row):
        header_like_cells = 0
        data_like_cells = 0
        non_empty_cells = 0
        for cell in row:
            if cell.value is not None:
                non_empty_cells += 1
                val = str(cell.value).strip()
                if self._looks_like_header(val):
                    header_like_cells += 1
                elif self._looks_like_data(val):
                    data_like_cells += 1
        if non_empty_cells == 0:
            return False
        return header_like_cells >= data_like_cells

    def _parse_simple_headers(self, rows):
        if not rows:
            return [], 0
        header_row = rows[0]
        headers = []
        for cell in header_row:
            if cell.value is not None:
                header_value = str(cell.value).strip()
                if header_value:
                    headers.append(header_value)
            else:
                pass
        final_headers = []
        for i, cell in enumerate(header_row):
            if cell.value is not None:
                header_value = str(cell.value).strip()
                if header_value:
                    final_headers.append(header_value)
                else:
                    final_headers.append(f"Column_{i + 1}")
            else:
                final_headers.append(f"Column_{i + 1}")
        return final_headers, 1

    def _parse_multi_level_headers(self, ws, rows):
        if len(rows) < 2:
            return [], 0
        header_rows = self._detect_header_rows(rows)
        if header_rows == 1:
            return self._parse_simple_headers(rows)
        else:
            return self._build_hierarchical_headers(ws, rows, header_rows), header_rows

    def _detect_header_rows(self, rows):
        if len(rows) < 2:
            return 1
        header_rows = 1
        max_check_rows = min(5, len(rows))
        for i in range(1, max_check_rows):
            row = rows[i]
            if self._row_looks_like_header(row):
                header_rows = i + 1
            else:
                break
        return header_rows

    def _looks_like_header(self, value):
        if len(value) < 1:
            return False
        if any(ord(c) > 127 for c in value):
            return True
        if len([c for c in value if c.isalpha()]) >= 2:
            return True
        if any(c in value for c in ["(", ")", "：", ":", "（", "）", "_", "-"]):
            return True
        return False

    def _looks_like_data(self, value):
        if len(value) == 1 and value.upper() in ["Y", "N", "M", "X", "/", "-"]:
            return True
        if value.replace(".", "").replace("-", "").replace(",", "").isdigit():
            return True
        if value.startswith("0x") and len(value) <= 10:
            return True
        return False

    def _build_hierarchical_headers(self, ws, rows, header_rows):
        headers = []
        max_col = max(len(row) for row in rows[:header_rows]) if header_rows > 0 else 0
        merged_ranges = list(ws.merged_cells.ranges)
        for col_idx in range(max_col):
            header_parts = []
            for row_idx in range(header_rows):
                if col_idx < len(rows[row_idx]):
                    cell_value = rows[row_idx][col_idx].value
                    merged_value = self._get_merged_cell_value(ws, row_idx + 1, col_idx + 1, merged_ranges)
                    if merged_value is not None:
                        cell_value = merged_value
                    if cell_value is not None:
                        cell_value = str(cell_value).strip()
                        if cell_value and cell_value not in header_parts and self._is_valid_header_part(cell_value):
                            header_parts.append(cell_value)
            if header_parts:
                header = "-".join(header_parts)
                headers.append(header)
            else:
                headers.append(f"Column_{col_idx + 1}")
        final_headers = [h for h in headers if h and h != "-"]
        return final_headers

    def _is_valid_header_part(self, value):
        if len(value) == 1 and value.upper() in ["Y", "N", "M", "X"]:
            return False
        if value.replace(".", "").replace("-", "").replace(",", "").isdigit():
            return False
        if value in ["/", "-", "+", "*", "="]:
            return False
        return True

    def _get_merged_cell_value(self, ws, row, col, merged_ranges):
        for merged_range in merged_ranges:
            if merged_range.min_row <= row <= merged_range.max_row and merged_range.min_col <= col <= merged_range.max_col:
                return ws.cell(merged_range.min_row, merged_range.min_col).value
        return None

    def _extract_row_data(self, ws, row, absolute_row_idx, expected_cols):
        row_data = []
        merged_ranges = list(ws.merged_cells.ranges)
        actual_row_num = absolute_row_idx + 1
        for col_idx in range(expected_cols):
            cell_value = None
            actual_col_num = col_idx + 1
            try:
                cell_value = ws.cell(row=actual_row_num, column=actual_col_num).value
            except ValueError:
                if col_idx < len(row):
                    cell_value = row[col_idx].value
            if cell_value is None:
                merged_value = self._get_merged_cell_value(ws, actual_row_num, actual_col_num, merged_ranges)
                if merged_value is not None:
                    cell_value = merged_value
                else:
                    cell_value = self._get_inherited_value(ws, actual_row_num, actual_col_num, merged_ranges)
            row_data.append(cell_value)
        return row_data

    def _get_inherited_value(self, ws, row, col, merged_ranges):
        for merged_range in merged_ranges:
            if merged_range.min_row <= row <= merged_range.max_row and merged_range.min_col <= col <= merged_range.max_col:
                return ws.cell(merged_range.min_row, merged_range.min_col).value
        return None

    def _is_empty_row(self, row_data):
        for val in row_data:
            if val is not None and str(val).strip() != "":
                return False
        return True
